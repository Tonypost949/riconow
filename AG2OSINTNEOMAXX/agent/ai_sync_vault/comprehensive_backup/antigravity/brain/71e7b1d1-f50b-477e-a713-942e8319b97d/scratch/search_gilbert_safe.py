import os
import openpyxl

scratch_dir = r"C:\Users\HP\.gemini\antigravity\brain\71e7b1d1-f50b-477e-a713-942e8319b97d\scratch"
output_file = os.path.join(scratch_dir, "gilbert_safe_results.txt")

search_terms = ["gilbert", "213", "685,000", "685000"]
files_to_search = ["osint_20260701T124651Z.xlsx", "osint_20260701T121450Z.xlsx"]

with open(output_file, "w", encoding="utf-8") as out:
    out.write("Safe Gilbert and 213 Search Results\n")
    out.write("===================================\n\n")
    
    for filename in files_to_search:
        filepath = os.path.join(scratch_dir, filename)
        if not os.path.exists(filepath):
            out.write(f"File not found: {filename}\n")
            continue
            
        out.write(f"Searching {filename}...\n")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                out.write(f"  Sheet: {sheetname}\n")
                row_count = 0
                for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    row_str = " | ".join([str(val) for val in row if val is not None])
                    row_str_lower = row_str.lower()
                    for term in search_terms:
                        if term in row_str_lower:
                            out.write(f"    Match (Row {r_idx}, Term '{term}'): {row_str}\n")
                            row_count += 1
                            break
                out.write(f"  Finished sheet {sheetname}. Matches found: {row_count}\n")
        except Exception as e:
            out.write(f"  Error reading {filename}: {str(e)}\n")
            
print("Done! Safe search results written to gilbert_safe_results.txt")
