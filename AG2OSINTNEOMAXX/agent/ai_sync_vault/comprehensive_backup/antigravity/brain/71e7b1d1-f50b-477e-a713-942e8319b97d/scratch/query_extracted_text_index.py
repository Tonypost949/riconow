import os
import openpyxl

scratch_dir = r"C:\Users\HP\.gemini\antigravity\brain\71e7b1d1-f50b-477e-a713-942e8319b97d\scratch"
file_path = os.path.join(scratch_dir, "osint_20260701T124651Z.xlsx")
output_file = os.path.join(scratch_dir, "extracted_text_index_filtered.txt")

print(f"Loading {file_path}...")
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

with open(output_file, "w", encoding="utf-8") as out:
    out.write("ExtractedTextIndex Sheet Filtered Search\n")
    out.write("========================================\n\n")
    
    if "ExtractedTextIndex" in wb.sheetnames:
        sheet = wb["ExtractedTextIndex"]
        out.write("Found sheet 'ExtractedTextIndex'. Scanning rows...\n")
        
        # Read header
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter)
        out.write(f"Header: {header}\n\n")
        
        match_count = 0
        for r_idx, row in enumerate(rows_iter, 2):
            row_str = " | ".join([str(val) for val in row if val is not None])
            row_lower = row_str.lower()
            
            # Check for gilbert, east, 685000, 685,000, or covenant
            if any(term in row_lower for term in ["gilbert", "east", "685000", "covenant"]):
                out.write(f"Row {r_idx}: {row_str}\n")
                match_count += 1
                
        out.write(f"\nScan completed. Found {match_count} matching rows in 'ExtractedTextIndex'.\n")
    else:
        out.write("Sheet 'ExtractedTextIndex' NOT found! Available sheets:\n")
        for name in wb.sheetnames:
            out.write(f"  - {name}\n")

print(f"Done! Filtered results written to {output_file}")
