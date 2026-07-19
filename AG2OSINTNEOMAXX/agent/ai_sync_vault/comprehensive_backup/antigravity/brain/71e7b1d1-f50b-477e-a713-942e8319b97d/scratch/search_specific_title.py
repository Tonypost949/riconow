import os
import openpyxl
import csv

scratch_dir = r"C:\Users\HP\.gemini\antigravity\brain\71e7b1d1-f50b-477e-a713-942e8319b97d\scratch"

search_terms = ["gilbert", "213", "685,000", "685000"]

print("Starting scan for terms:", search_terms)

# 1. Search text/csv files
for root, dirs, files in os.walk(scratch_dir):
    for filename in files:
        filepath = os.path.join(root, filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext in [".txt", ".csv", ".json", ".md"]:
            try:
                with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                    for i, line in enumerate(f, 1):
                        line_lower = line.lower()
                        for term in search_terms:
                            if term in line_lower:
                                # print match context
                                print(f"[TXT] {filename}:{i} (term '{term}'): {line.strip()[:150]}")
            except Exception as e:
                pass

        elif ext in [".xlsx"]:
            try:
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                for sheetname in wb.sheetnames:
                    sheet = wb[sheetname]
                    # We iterate rows
                    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                        for c_idx, val in enumerate(row, 1):
                            if val is not None:
                                val_str = str(val).lower()
                                for term in search_terms:
                                    if term in val_str:
                                        print(f"[XLSX] {filename} | Sheet: {sheetname} | Row {r_idx}, Col {c_idx} (term '{term}'): {val}")
            except Exception as e:
                print(f"Error reading Excel {filename}: {e}")
